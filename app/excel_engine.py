import pandas as pd
from openpyxl import load_workbook
from langchain_openai import ChatOpenAI
from pypdf import PdfReader
from app.config import settings

def extract_text_from_pdf(file_path: str):
    reader = PdfReader(file_path)
    text = ""
    for page in reader.pages:
        text += page.extract_text()
    return text

def fill_excel_from_pdf(excel_template_path: str, pdf_path: str, output_path: str):
    try:
        # Load Excel template
        wb = load_workbook(excel_template_path)
        ws = wb.active

        # Read headers from Excel
        headers = [cell.value for cell in ws[1] if cell.value]

        # Extract text from PDF
        pdf_text = extract_text_from_pdf(pdf_path)

        # Use LLM to generate JSON based on headers and PDF text
        llm = ChatOpenAI(model=settings.default_llm, openai_api_key=settings.openai_api_key)
        prompt = f"Extract the following fields from the text: {', '.join(headers)}. Text: {pdf_text}. Return as JSON."
        response = llm.predict(prompt)

        # Parse response as JSON (assuming it's valid JSON)
        import json
        data = json.loads(response)

        # Fill Excel
        for col_num, header in enumerate(headers, 1):
            if header in data:
                ws.cell(row=2, column=col_num, value=data[header])

        # Save
        wb.save(output_path)
        return {"message": "Excel filled and saved"}
    except Exception as e:
        return {"error": str(e)}
