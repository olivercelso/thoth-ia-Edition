import os

# Function to create directories
def create_directories():
    os.makedirs('app', exist_ok=True)

# Function to write files
def write_file(file_path, content):
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(content)

# Content for each file
files_content = {
    'requirements.txt': """fastapi
uvicorn
python-multipart
sqlalchemy
chromadb
langchain
langchain-openai
pandas
openpyxl
pypdf
slowapi
python-dotenv
""",
    '.env.example': """OPENAI_API_KEY=
ADMIN_API_KEY=
ALLOWED_ORIGINS=
""",
    'app/config.py': """from pydantic_settings import BaseSettings
from typing import List

class Settings(BaseSettings):
    openai_api_key: str
    admin_api_key: str
    allowed_origins: List[str] = ["http://localhost:8000"]
    default_llm: str = "gpt-4o-mini"
    vision_llm: str = "gpt-4o"

    class Config:
        env_file = ".env"

settings = Settings()
""",
    'app/security.py': """from fastapi import HTTPException, Depends, Security
from fastapi.security.api_key import APIKeyHeader
from app.config import settings

api_key_header = APIKeyHeader(name="X-API-Key", auto_error=False)

def get_api_key(api_key: str = Security(api_key_header)):
    if api_key != settings.admin_api_key:
        raise HTTPException(status_code=401, detail="Invalid API Key")
    return api_key
""",
    'app/database.py': """from sqlalchemy import create_engine
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker

SQLALCHEMY_DATABASE_URL = "sqlite:///./neuralmanager.db"

engine = create_engine(SQLALCHEMY_DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

Base = declarative_base()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()
""",
    'app/models.py': """from sqlalchemy import Column, Integer, String, DateTime, Text
from sqlalchemy.sql import func
from app.database import Base

class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, unique=True, index=True)
    email = Column(String, unique=True, index=True)
    created_at = Column(DateTime(timezone=True), server_default=func.now())

class Document(Base):
    __tablename__ = "documents"
    id = Column(Integer, primary_key=True, index=True)
    filename = Column(String)
    content = Column(Text)
    created_at = Column(DateTime(timezone=True), server_default=func.now())

class AuditLog(Base):
    __tablename__ = "audit_logs"
    id = Column(Integer, primary_key=True, index=True)
    action = Column(String)
    user_id = Column(Integer)
    timestamp = Column(DateTime(timezone=True), server_default=func.now())
""",
    'app/rag_module.py': """import chromadb
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_openai import OpenAIEmbeddings, ChatOpenAI
from langchain.vectorstores import Chroma
from langchain.chains import RetrievalQA
from pypdf import PdfReader
from app.config import settings

# Initialize ChromaDB
client = chromadb.PersistentClient(path="./chroma_db")
collection = client.get_or_create_collection(name="documents")

embeddings = OpenAIEmbeddings(openai_api_key=settings.openai_api_key)

def ingest_pdf(file_path: str):
    try:
        reader = PdfReader(file_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text()

        text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=200)
        chunks = text_splitter.split_text(text)

        # Add to ChromaDB
        ids = [f"chunk_{i}" for i in range(len(chunks))]
        collection.add(documents=chunks, ids=ids)

        # Also add to LangChain Chroma for retrieval
        vectorstore = Chroma.from_documents(
            documents=[{"page_content": chunk, "metadata": {}} for chunk in chunks],
            embedding=embeddings,
            persist_directory="./chroma_db"
        )
        vectorstore.persist()

        return {"message": "PDF ingested successfully"}
    except Exception as e:
        return {"error": str(e)}

def ask_rag(question: str):
    try:
        llm = ChatOpenAI(model=settings.default_llm, openai_api_key=settings.openai_api_key)
        vectorstore = Chroma(persist_directory="./chroma_db", embedding_function=embeddings)
        qa_chain = RetrievalQA.from_chain_type(llm=llm, chain_type="stuff", retriever=vectorstore.as_retriever())
        answer = qa_chain.run(question)
        return {"answer": answer}
    except Exception as e:
        return {"error": str(e)}
""",
    'app/vision_module.py': """import base64
from langchain_openai import ChatOpenAI
from langchain.schema import HumanMessage
from app.config import settings

def analyze_image(image_bytes: bytes, prompt: str = "Describe this image"):
    try:
        llm = ChatOpenAI(model=settings.vision_llm, openai_api_key=settings.openai_api_key)

        # Convert bytes to base64
        image_b64 = base64.b64encode(image_bytes).decode('utf-8')
        image_url = f"data:image/jpeg;base64,{image_b64}"

        message = HumanMessage(
            content=[
                {"type": "text", "text": prompt},
                {"type": "image_url", "image_url": {"url": image_url}}
            ]
        )

        response = llm([message])
        return {"analysis": response.content}
    except Exception as e:
        return {"error": str(e)}
""",
    'app/excel_engine.py': """import pandas as pd
from openpyxl import load_workbook
from langchain_openai import ChatOpenAI
from pypdf import PdfReader
from app.config import settings

def extract_text_from_pdf(file_path: str):
    reader = PdfReader(file_path)
    text = ""
    for page in reader.pages:
        text += page.extract_text()
    return text

def fill_excel_from_pdf(excel_template_path: str, pdf_path: str, output_path: str):
    try:
        # Load Excel template
        wb = load_workbook(excel_template_path)
        ws = wb.active

        # Read headers from Excel
        headers = [cell.value for cell in ws[1] if cell.value]

        # Extract text from PDF
        pdf_text = extract_text_from_pdf(pdf_path)

        # Use LLM to generate JSON based on headers and PDF text
        llm = ChatOpenAI(model=settings.default_llm, openai_api_key=settings.openai_api_key)
        prompt = f"Extract the following fields from the text: {', '.join(headers)}. Text: {pdf_text}. Return as JSON."
        response = llm.predict(prompt)

        # Parse response as JSON (assuming it's valid JSON)
        import json
        data = json.loads(response)

        # Fill Excel
        for col_num, header in enumerate(headers, 1):
            if header in data:
                ws.cell(row=2, column=col_num, value=data[header])

        # Save
        wb.save(output_path)
        return {"message": "Excel filled and saved"}
    except Exception as e:
        return {"error": str(e)}
""",
    'app/main.py': """from fastapi import FastAPI, UploadFile, File, Depends, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from slowapi.middleware import SlowAPIMiddleware
from app.config import settings
from app.security import get_api_key
from app.rag_module import ingest_pdf, ask_rag
from app.vision_module import analyze_image
from app.excel_engine import fill_excel_from_pdf
from app.database import engine, Base

# Create tables
Base.metadata.create_all(bind=engine)

app = FastAPI()

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Rate Limiting
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
app.add_middleware(SlowAPIMiddleware)

@app.post("/chat")
@limiter.limit("20/minute")
def chat(question: str, api_key: str = Depends(get_api_key)):
    return ask_rag(question)

@app.post("/vision")
@limiter.limit("20/minute")
def vision(file: UploadFile = File(...), prompt: str = "Describe this image", api_key: str = Depends(get_api_key)):
    try:
        image_bytes = file.file.read()
        return analyze_image(image_bytes, prompt)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/fill-excel")
@limiter.limit("20/minute")
def fill_excel(excel_file: UploadFile = File(...), pdf_file: UploadFile = File(...), api_key: str = Depends(get_api_key)):
    try:
        # Save temporary files
        with open("temp_template.xlsx", "wb") as f:
            f.write(excel_file.file.read())
        with open("temp_pdf.pdf", "wb") as f:
            f.write(pdf_file.file.read())

        output_path = "filled_excel.xlsx"
        result = fill_excel_from_pdf("temp_template.xlsx", "temp_pdf.pdf", output_path)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
"""
}

if __name__ == "__main__":
    create_directories()
    for file_path, content in files_content.items():
        write_file(file_path, content)
    print("Projeto NeuralManager criado com sucesso! Execute: pip install -r requirements.txt")
